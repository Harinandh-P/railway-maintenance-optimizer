import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelService:
    @staticmethod
    def export_to_excel(data: list, sheet_name: str = "Data") -> bytes:
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Style header row
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-fit column width
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output.seek(0)
        return output.getvalue()
